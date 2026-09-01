"""
excel/parse_excel.py
Usage:  python parse_excel.py <filename>
Place file in input/ folder. Output goes to output/ with a datetime stamp.
Supports: .xlsx  .xlsm  .xltx  .xls  .csv
"""

import sys, os
from datetime import datetime

def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_excel.py <filename>")
        print("Example: python parse_excel.py data.xlsx")
        sys.exit(1)

    filename   = sys.argv[1].strip()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, "input", filename)

    if not os.path.isfile(input_path):
        print(f"ERROR: File not found: {input_path}")
        print(f"Place '{filename}' inside the input/ folder and try again.")
        sys.exit(1)

    ext = os.path.splitext(filename)[1].lower()

    # ── install dependencies if missing ────────────────────────────────────
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not found. Installing...")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    try:
        import pandas as pd
    except ImportError:
        print("pandas not found. Installing...")
        os.system("pip install pandas --break-system-packages -q")
        import pandas as pd

    lines = []
    lines.append(f"FILE: {filename}")
    lines.append(f"PARSED AT: {datetime.now()}")
    lines.append(f"EXTENSION: {ext}")

    # ── xlsx / xlsm / xltx ────────────────────────────────────────────────
    if ext in (".xlsx", ".xlsm", ".xltx"):
        try:
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"ERROR opening file: {e}")
            sys.exit(1)

        # workbook-level properties
        lines.append("\n" + "="*60)
        lines.append("WORKBOOK PROPERTIES")
        lines.append("="*60)
        try:
            props = wb.properties
            for attr in ["title","subject","creator","keywords","description",
                         "lastModifiedBy","created","modified","category"]:
                try:
                    lines.append(f"{attr}: {getattr(props, attr, None)}")
                except Exception as e:
                    lines.append(f"{attr}: [ERROR: {e}]")
        except Exception as e:
            lines.append(f"[Could not read properties: {e}]")

        try:
            lines.append(f"\nsheets: {wb.sheetnames}")
        except Exception as e:
            lines.append(f"sheets: [ERROR: {e}]")

        # every sheet, every row, every cell — raw dump
        try:
            for sheet_name in wb.sheetnames:
                lines.append("\n" + "="*60)
                lines.append(f"SHEET: {sheet_name}")
                lines.append("="*60)
                try:
                    ws = wb[sheet_name]
                    try:
                        lines.append(f"dimensions: {ws.dimensions}")
                    except Exception as e:
                        lines.append(f"dimensions: [ERROR: {e}]")
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        try:
                            cells = [str(c) if c is not None else "" for c in row]
                            lines.append(f"row[{row_idx}]: {' | '.join(cells)}")
                        except Exception as e:
                            lines.append(f"row[{row_idx}]: [ERROR: {e}]")
                except Exception as e:
                    lines.append(f"[Could not read sheet '{sheet_name}': {e}]")
        except Exception as e:
            lines.append(f"[Could not iterate sheets: {e}]")

        try:
            wb.close()
        except Exception:
            pass

    # ── legacy .xls ───────────────────────────────────────────────────────
    elif ext == ".xls":
        try:
            import xlrd
        except ImportError:
            print("xlrd not found. Installing...")
            os.system("pip install xlrd --break-system-packages -q")
            import xlrd

        try:
            wb = xlrd.open_workbook(input_path)
        except Exception as e:
            print(f"ERROR opening file: {e}")
            sys.exit(1)

        try:
            lines.append(f"\nsheets: {wb.sheet_names()}")
        except Exception as e:
            lines.append(f"sheets: [ERROR: {e}]")

        try:
            for sheet_name in wb.sheet_names():
                lines.append("\n" + "="*60)
                lines.append(f"SHEET: {sheet_name}")
                lines.append("="*60)
                try:
                    ws = wb.sheet_by_name(sheet_name)
                    lines.append(f"rows: {ws.nrows}  cols: {ws.ncols}")
                    for row_idx in range(ws.nrows):
                        try:
                            cells = [str(ws.cell_value(row_idx, col_idx))
                                     for col_idx in range(ws.ncols)]
                            lines.append(f"row[{row_idx}]: {' | '.join(cells)}")
                        except Exception as e:
                            lines.append(f"row[{row_idx}]: [ERROR: {e}]")
                except Exception as e:
                    lines.append(f"[Could not read sheet '{sheet_name}': {e}]")
        except Exception as e:
            lines.append(f"[Could not iterate sheets: {e}]")

    # ── csv ───────────────────────────────────────────────────────────────
    elif ext == ".csv":
        lines.append("\n" + "="*60)
        lines.append("CSV CONTENT")
        lines.append("="*60)
        try:
            with open(input_path, "r", encoding="utf-8", errors="replace") as f:
                for row_idx, raw_line in enumerate(f):
                    lines.append(f"row[{row_idx}]: {raw_line.rstrip()}")
        except Exception as e:
            lines.append(f"[Could not read CSV: {e}]")

    else:
        lines.append(f"[Unsupported extension: {ext}. Supported: .xlsx .xlsm .xltx .xls .csv]")

    # ── write output ───────────────────────────────────────────────────────
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    base       = os.path.splitext(filename)[0]
    output_dir = os.path.join(script_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{base}__{timestamp}.txt")

    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        print(f"Done! Output → {output_path}")
    except Exception as e:
        print(f"ERROR writing output: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
